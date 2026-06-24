#!/usr/bin/env python3
"""
build_template.py — generate the matching BLANK Excel template
(al-ahli-sijillat-template.xlsx) from the SAME data.js the web form renders.

Guarantees the spreadsheet form and the web form share one schema (no drift):
it shells out to node to dump REGISTER -> JSON, then builds the workbook with
openpyxl: RTL sheet view, Arabic headers, one worksheet per program, year
columns mirroring the register, plus document/PII status + count + notes columns.

Run:  python3 build_template.py
Out:  al-ahli-sijillat-template.xlsx  (next to this script)
"""
import json
import os
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_JS = os.path.join(HERE, "data.js")
OUT = os.path.join(HERE, "al-ahli-sijillat-template.xlsx")

SHAPE_LABEL = {
    "matrix": "مصفوفة سنوات",
    "document": "وثيقة مرفقة",
    "select": "حالة (اختيار)",
    "pii": "قائمة مستفيدين (PII)",
}
AVAIL_LABEL = {
    "available": "متوفر الآن ✓",
    "fetch": "يحتاج جلب",
    "permission": "يحتاج إذن / إحالة",
}
STAGES = ["مدخلات", "أنشطة", "مخرجات", "نتائج", "أثر"]

# --- palette (mirrors the web form) -----------------------------------------
SNB_DARK = "0F5240"
SNB_MID = "156552"
SAND_HEAD = "F3EFE7"
STAGE_FILL = "EEF6F2"
UNREV_FILL = "ECE4F7"
WHITE = "FFFFFF"

thin = Side(style="thin", color="D3CAB7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def load_register():
    """Dump REGISTER from data.js using node, parse as JSON."""
    script = (
        "const fs=require('fs');"
        "global.window={};"
        f"eval(fs.readFileSync({json.dumps(DATA_JS)},'utf8'));"
        "process.stdout.write(JSON.stringify(window.REGISTER));"
    )
    out = subprocess.check_output(["node", "-e", script])
    return json.loads(out)


def item_years(item, program):
    return item.get("years") or program["years"]


def build():
    reg = load_register()
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(name="Arial", size=13, bold=True, color=SNB_DARK)
    meta_font = Font(name="Arial", size=10, color="3C463F")
    head_font = Font(name="Arial", size=10, bold=True, color="0A3D2E")
    cell_font = Font(name="Arial", size=10, color="1C241F")
    warn_font = Font(name="Arial", size=10, bold=True, color="6B4EA8")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for idx, p in enumerate(reg["programs"], start=1):
        name = p["name"]
        safe = f"{idx}-{name}"
        for ch in '\\/?*[]:':
            safe = safe.replace(ch, "")
        ws = wb.create_sheet(title=safe[:31])
        ws.sheet_view.rightToLeft = True

        # union of all years (program + item overrides)
        years = list(p["years"])
        for it in p["items"]:
            for y in (it.get("years") or []):
                if y not in years:
                    years.append(y)
        years.sort()

        r = 1
        ws.cell(r, 1, f"البرنامج {idx}: {name}").font = title_font
        r += 1
        ws.cell(r, 1, "السنوات: " + " – ".join(str(y) for y in p["years"])).font = meta_font
        r += 1
        ws.cell(r, 1, p.get("note", "")).font = meta_font
        ws.cell(r, 1).alignment = right
        r += 2

        # header row
        headers = ["المرحلة", "البند / المؤشر", "شكل البيان", "الحالة (التوافر)", "المصدر / المالك"]
        headers += [str(y) for y in years]
        headers += ["حالة المستند/القائمة", "العدد", "ملاحظات"]
        head_row = r
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(r, ci, h)
            c.font = head_font
            c.fill = PatternFill("solid", fgColor=SAND_HEAD)
            c.alignment = center
            c.border = BORDER
        r += 1

        for stage in STAGES:
            stage_items = [it for it in p["items"] if it["stage"] == stage]
            if not stage_items:
                continue
            for it in stage_items:
                avail = AVAIL_LABEL.get(it.get("status"), "")
                vals = [it["stage"], it["label"], SHAPE_LABEL[it["shape"]], avail, it.get("source", "")]
                iy = item_years(it, p)
                for y in years:
                    # blank template: cells left empty; mark N/A where year not in scope
                    if it["shape"] == "matrix" and y in iy:
                        vals.append("")
                    else:
                        vals.append("—" if it["shape"] == "matrix" else "")
                # status / count / notes (blank for client to fill)
                vals += ["", "", ""]
                for ci, v in enumerate(vals, start=1):
                    c = ws.cell(r, ci, v)
                    c.font = cell_font
                    c.border = BORDER
                    c.alignment = right if ci in (1, 2, 5) else center
                r += 1

        # column widths
        widths = [10, 48, 16, 18, 24] + [10] * len(years) + [26, 8, 30]
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[head_row].height = 26
        ws.freeze_panes = ws.cell(head_row + 1, 1)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"programs: {len(reg['programs'])}, total items: {sum(len(p['items']) for p in reg['programs'])}")


if __name__ == "__main__":
    try:
        build()
    except subprocess.CalledProcessError as e:
        print("node failed to parse data.js:", e, file=sys.stderr)
        sys.exit(1)
